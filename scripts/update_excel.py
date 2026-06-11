"""
update_excel.py
===============
Writes the draw results from participants.json into the
'Full Results' sheet of the staff list Excel file.
"""

import json, pathlib, shutil
from openpyxl import load_workbook

ROOT         = pathlib.Path(__file__).parent.parent
PARTICIPANTS = ROOT / 'data' / 'participants.json'
EXCEL_IN     = pathlib.Path(r'C:\Users\AngusGorrie\Downloads\staff list for sweepstake.xlsx')
EXCEL_OUT    = pathlib.Path(r'C:\Users\AngusGorrie\Downloads\staff list for sweepstake (with draw).xlsx')

with open(PARTICIPANTS, encoding='utf-8') as f:
    participants = json.load(f)

# Build lookup: name -> (front_runner, long_shot, not_a_chancer)
draw = {}
for p in participants:
    teams = {t['bracket']: t['name'] for t in p['teams']}
    draw[p['name']] = (
        teams.get('front-runner', ''),
        teams.get('long-shot', ''),
        teams.get('not-a-chancer', ''),
    )

wb = load_workbook(EXCEL_IN)
ws = wb['Drawn teams']

# Find header row to locate Name, Front Runner, Long Shot, Not A Chance columns
header_row = None
col_name = col_fr = col_ls = col_nac = None

for row in ws.iter_rows():
    for cell in row:
        val = str(cell.value).strip() if cell.value else ''
        if val == 'Name':
            header_row = cell.row
            col_name = cell.column
        elif val in ('Front Runner', 'Front-Runner', 'Front Runner '):
            col_fr = cell.column
        elif val in ('Long Shot', 'Long-Shot', 'Long Shot '):
            col_ls = cell.column
        elif val in ('Not A Chance', 'Not-A-Chancer', 'Not A Chance ', 'Not-A-Chance'):
            col_nac = cell.column

print(f"Header row: {header_row}, Name col: {col_name}, FR col: {col_fr}, LS col: {col_ls}, NAC col: {col_nac}")

matched = 0
unmatched = []

for row in ws.iter_rows(min_row=header_row + 1):
    name_cell = None
    for cell in row:
        if cell.column == col_name:
            name_cell = cell
            break
    if not name_cell or not name_cell.value:
        continue

    name = str(name_cell.value).strip()
    if name in draw:
        fr, ls, nac = draw[name]
        ws.cell(row=name_cell.row, column=col_fr).value = fr
        ws.cell(row=name_cell.row, column=col_ls).value = ls
        ws.cell(row=name_cell.row, column=col_nac).value = nac
        matched += 1
    else:
        unmatched.append(name)

print(f"Updated {matched} participants.")
if unmatched:
    print(f"No match found for: {unmatched}")

wb.save(EXCEL_OUT)
print(f"Saved to {EXCEL_OUT}")
